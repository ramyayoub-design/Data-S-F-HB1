"""
Speckle Automate Function
Extracts BrepX properties from 'plugins', 'Core HB1-blockA', 'filtration', and 'pollution' collections
and exports them to Excel + Google Sheets.
"""

import json
import sys
import traceback
import threading
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from enum import Enum
from typing import Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError as e:
    print(f"IMPORT ERROR: {e}", flush=True)
    traceback.print_exc()
    sys.exit(1)

from pydantic import Field
from speckle_automate import AutomateBase, AutomationContext, execute_automate_function
from specklepy.objects.base import Base
from flatten import flatten_base


# ─── Email configuration ──────────────────────────────────────────────────────

SENDER_EMAIL    = "ramy.ayoub@students.iaac.net"      # ← replace with the sending Gmail address
SENDER_PASSWORD = "ufmb bngr rakd qidd"        # ← replace with the Gmail App Password
TEAM_EMAILS     = [
    "maria.sanchez.i.dominguez@students.iaac.net",
    "charles.abi.chahine@students.iaac.net",
    "lakzhmy.mari.zaro@students.iaac.net",
    "emilie.elchidiac@students.iaac.net",
    "hani.karime@students.iaac.net",


    # add more teammates below:
    # "teammate2@students.iaac.net",
    # "teammate3@students.iaac.net",
]


# ─── Inputs ───────────────────────────────────────────────────────────────────

class OutputFormat(str, Enum):
    EXCEL_ONLY  = "excel_only"
    SHEETS_ONLY = "sheets_only"
    BOTH        = "both"


class FunctionInputs(AutomateBase):
    output_format: OutputFormat = Field(
        default=OutputFormat.BOTH,
        title="Output Format",
        description="Choose where to export the extracted data.",
    )
    google_sheet_id: str = Field(
        title="Google Sheet ID",
        description="The ID from your Sheet URL: .../spreadsheets/d/<ID>/edit",
        min_length=10,
    )
    google_service_account_json: str = Field(
        title="Google Service Account JSON",
        description="Full JSON content of your GCP service account key.",
    )


# ─── Helpers ──────────────────────────────────────────────────────────────────

def get_prop(obj: Base, *key_fragments: str) -> Any:
    props = getattr(obj, "properties", None)
    if not props:
        return None
    prop_dict = (
        {k: getattr(props, k) for k in props.get_dynamic_member_names()}
        if isinstance(props, Base)
        else props if isinstance(props, dict)
        else None
    )
    if not prop_dict:
        return None
    for fragment in key_fragments:
        for k, v in prop_dict.items():
            if fragment.lower() in k.lower().strip():
                return v
    return None


def style_header_row(ws, row: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_data_row(ws, row: int, color_a: str, color_b: str, is_even: bool):
    fill = PatternFill("solid", fgColor=color_b if is_even else color_a)
    for cell in ws[row]:
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_total_row(ws, row: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ─── Sheet builders ───────────────────────────────────────────────────────────

def build_plugins_sheet(ws, breps: List[Base]):
    # Colors: deep orange header, alternating light/mid orange rows
    HEADER_COLOR = "BF4B04"
    ROW_COLOR_A  = "FFD8B0"   # light orange
    ROW_COLOR_B  = "FFC080"   # mid orange
    TOTAL_COLOR  = "BF4B04"

    headers = [
        "Index", "Speckle ID", "Application ID",
        "Geometry Area (m²)", "Geometry Volume (m³)",
        "Volume (brep #)", "Normalized Score",
        "STR_PAR_Density", "ENV_PAR_WindPressure", "ENV_PAR_IncidentRadiation",
    ]
    ws.append(headers)
    style_header_row(ws, 1, HEADER_COLOR)

    for i, brep in enumerate(breps, start=1):
        ws.append([
            i,
            getattr(brep, "id", None),
            getattr(brep, "applicationId", None),
            round(getattr(brep, "area",   0) or 0, 4),
            round(getattr(brep, "volume", 0) or 0, 4),
            get_prop(brep, "Volume"),
            get_prop(brep, "Normalized"),
            get_prop(brep, "STR_PAR_Density"),
            get_prop(brep, "ENV_PAR_WindPressure"),
            get_prop(brep, "ENV_PAR_IncidentRadiation"),
        ])
        style_data_row(ws, i + 1, ROW_COLOR_A, ROW_COLOR_B, i % 2 == 0)

    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["TOTAL BREPS", len(breps)])
    style_total_row(ws, ws.max_row, TOTAL_COLOR)
    autofit_columns(ws)


def build_core_sheet(ws, breps: List[Base]):
    # Colors: deep green header, alternating light/mid green rows
    HEADER_COLOR = "1E5631"
    ROW_COLOR_A  = "C8E6C9"
    ROW_COLOR_B  = "A5D6A7"
    TOTAL_COLOR  = "1E5631"

    headers = [
        "Index", "Speckle ID", "Application ID",
        "Stress Pts Coordinates", "Beam Thickness (m)", "STR_PAR_StressLoad",
    ]
    ws.append(headers)
    style_header_row(ws, 1, HEADER_COLOR)

    total_length = 0.0
    for i, brep in enumerate(breps, start=1):
        stress = get_prop(brep, "Stress", "stress pts", "stress_pts")
        beam   = get_prop(brep, "Beam", "thickness", "beam thick")
        length = getattr(brep, "length", None)
        if length:
            try:
                total_length += float(length)
            except (TypeError, ValueError):
                pass
        ws.append([
            i,
            getattr(brep, "id", None),
            getattr(brep, "applicationId", None),
            str(stress) if stress is not None else None,
            beam,
            get_prop(brep, "STR_PAR_StressLoad"),
        ])
        style_data_row(ws, i + 1, ROW_COLOR_A, ROW_COLOR_B, i % 2 == 0)

    ws.append([])
    ws.append(["TOTAL BREPS", len(breps)])
    style_total_row(ws, ws.max_row, TOTAL_COLOR)
    if total_length > 0:
        ws.append(["TOTAL LENGTH (m)", round(total_length, 4)])
        style_total_row(ws, ws.max_row, TOTAL_COLOR)
    autofit_columns(ws)


def build_filtration_sheet(ws, breps: List[Base]):
    HEADER_COLOR = "4A148C"
    ROW_COLOR_A  = "E1BEE7"
    ROW_COLOR_B  = "CE93D8"
    TOTAL_COLOR  = "4A148C"

    headers = [
        "Index", "Speckle ID", "Application ID",
        "STR_PAR_FiltrationEfficiency",
    ]
    ws.append(headers)
    style_header_row(ws, 1, HEADER_COLOR)

    for i, brep in enumerate(breps, start=1):
        ws.append([
            i,
            getattr(brep, "id", None),
            getattr(brep, "applicationId", None),
            get_prop(brep, "STR_PAR_FiltrationEfficiency"),
        ])
        style_data_row(ws, i + 1, ROW_COLOR_A, ROW_COLOR_B, i % 2 == 0)

    ws.append([])
    ws.append(["TOTAL BREPS", len(breps)])
    style_total_row(ws, ws.max_row, TOTAL_COLOR)
    autofit_columns(ws)


def build_pollution_sheet(ws, breps: List[Base]):
    HEADER_COLOR = "7B3F00"
    ROW_COLOR_A  = "FFE0B2"
    ROW_COLOR_B  = "FFCC80"
    TOTAL_COLOR  = "7B3F00"

    headers = [
        "Index", "Speckle ID", "Application ID",
        "ENV_PAR_ExternalPollution",
    ]
    ws.append(headers)
    style_header_row(ws, 1, HEADER_COLOR)

    for i, brep in enumerate(breps, start=1):
        ws.append([
            i,
            getattr(brep, "id", None),
            getattr(brep, "applicationId", None),
            get_prop(brep, "ENV_PAR_ExternalPollution"),
        ])
        style_data_row(ws, i + 1, ROW_COLOR_A, ROW_COLOR_B, i % 2 == 0)

    ws.append([])
    ws.append(["TOTAL BREPS", len(breps)])
    style_total_row(ws, ws.max_row, TOTAL_COLOR)
    autofit_columns(ws)


# ─── Email notification ───────────────────────────────────────────────────────

def send_email_notification(sheet_url: str, speckle_url: str) -> str:
    if not TEAM_EMAILS:
        return "Email skipped: no recipient addresses in TEAM_EMAILS."

    subject = "Speckle Automate run complete — Data S-F HB1"
    body = f"""\
Hi team,

The latest Speckle Automate run for project Data S-F HB1 has finished successfully.

Please find the results here:

  Google Sheet:
  {sheet_url}

  Speckle Model:
  {speckle_url}

This message was sent automatically by Speckle Automate.
"""

    msg = MIMEMultipart()
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = ", ".join(TEAM_EMAILS)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, TEAM_EMAILS, msg.as_string())

    return f"Email sent to: {', '.join(TEAM_EMAILS)}"


# ─── Google Sheets sync ───────────────────────────────────────────────────────

def sync_to_google_sheets(sheet_id: str, service_account_json: str, wb: openpyxl.Workbook):
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(json.loads(service_account_json), scopes=scopes)
    gc = gspread.authorize(creds)
    spreadsheet = gc.open_by_key(sheet_id)

    # Delete default Sheet1 if it exists
    for default_name in ["Sheet1", "sheet1", "Feuille 1", "Hoja 1"]:
        try:
            default_ws = spreadsheet.worksheet(default_name)
            # Only delete if it's not the last sheet
            if len(spreadsheet.worksheets()) > 1:
                spreadsheet.del_worksheet(default_ws)
        except gspread.WorksheetNotFound:
            pass

    for sheet_name in wb.sheetnames:
        rows = [[cell.value for cell in row] for row in wb[sheet_name].iter_rows()]
        try:
            gs_ws = spreadsheet.worksheet(sheet_name)
            gs_ws.clear()
        except gspread.WorksheetNotFound:
            gs_ws = spreadsheet.add_worksheet(title=sheet_name, rows=200, cols=30)
        if rows:
            gs_ws.update(rows, value_input_option="USER_ENTERED")


# ─── Main function ────────────────────────────────────────────────────────────

def automate_function(
    automate_context: AutomationContext,
    function_inputs: FunctionInputs,
) -> None:
    try:
        print("DEBUG: function started", flush=True)

        # Timeout wrapper around receive_version()
        result_holder = [None]
        error_holder  = [None]

        def do_receive():
            try:
                result_holder[0] = automate_context.receive_version()
            except Exception as e:
                error_holder[0] = e

        t = threading.Thread(target=do_receive)
        t.start()
        t.join(timeout=50)

        if t.is_alive():
            print("TIMEOUT: receive_version() hung for 50s", flush=True)
            automate_context.mark_run_failed("receive_version() timed out")
            return

        if error_holder[0]:
            print(f"ERROR in receive_version: {error_holder[0]}", flush=True)
            traceback.print_exc()
            automate_context.mark_run_failed(str(error_holder[0]))
            return

        version_root_object = result_holder[0]
        print("DEBUG: version received", flush=True)

        # Get all BrepX objects
        all_breps = [e for e in flatten_base(version_root_object) if "Brep" in getattr(e, "speckle_type", "")]
        print(f"DEBUG: total breps={len(all_breps)}", flush=True)

        # Split by property keys
        plugin_breps     = []
        core_breps       = []
        filtration_breps = []
        pollution_breps  = []

        for brep in all_breps:
            props = getattr(brep, "properties", None)
            if props:
                keys = list(props.get_dynamic_member_names()) if isinstance(props, Base) else list(props.keys()) if isinstance(props, dict) else []
                keys_lower = [k.lower().strip() for k in keys]
                if any("str_par_filtrationefficiency" in k for k in keys_lower):
                    filtration_breps.append(brep)
                elif any("env_par_externalpollution" in k for k in keys_lower):
                    pollution_breps.append(brep)
                elif any("volume" in k for k in keys_lower):
                    plugin_breps.append(brep)
                else:
                    core_breps.append(brep)
            else:
                core_breps.append(brep)

        print(f"DEBUG: plugins={len(plugin_breps)} core={len(core_breps)} filtration={len(filtration_breps)} pollution={len(pollution_breps)}", flush=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        build_plugins_sheet(   wb.create_sheet("Plugins - Volumes"),      plugin_breps)
        build_core_sheet(      wb.create_sheet("Core HB1-blockA"),        core_breps)
        build_filtration_sheet(wb.create_sheet("Filtration"),             filtration_breps)
        build_pollution_sheet( wb.create_sheet("Pollution"),              pollution_breps)

        xlsx_path     = "/tmp/speckle_export.xlsx"
        sheets_status = "Google Sheets skipped."


        if function_inputs.output_format in (OutputFormat.EXCEL_ONLY, OutputFormat.BOTH):
            wb.save(xlsx_path)
            print("DEBUG: excel saved locally", flush=True)
            try:
                automate_context.store_file_result(xlsx_path)
                print("DEBUG: excel uploaded to Speckle", flush=True)
            except Exception as e:
                print(f"DEBUG: excel upload failed (file too large): {e}", flush=True)

        json_val = function_inputs.google_service_account_json
        print(f"DEBUG: json field length={len(json_val)} first50={json_val[:50]!r}", flush=True)

        if function_inputs.output_format in (OutputFormat.SHEETS_ONLY, OutputFormat.BOTH):
            try:
                sync_to_google_sheets(
                    sheet_id=function_inputs.google_sheet_id,
                    service_account_json=function_inputs.google_service_account_json,
                    wb=wb,
                )
                sheets_status = "Google Sheets synced."
            except Exception as e:
                sheets_status = f"Google Sheets sync failed: {e}"
            print(f"DEBUG: {sheets_status}", flush=True)

        # Build URLs for the notification email
        run_data     = automate_context.automation_run_data
        server_url   = run_data.speckle_server_url.rstrip("/")
        project_id   = run_data.project_id
        model_id     = run_data.model_id
        version_id   = run_data.version_id
        speckle_url  = f"{server_url}/projects/{project_id}/models/{model_id}@{version_id}"
        sheet_url    = f"https://docs.google.com/spreadsheets/d/{function_inputs.google_sheet_id}/edit"

        email_status = "Email skipped."
        try:
            email_status = send_email_notification(sheet_url=sheet_url, speckle_url=speckle_url)
        except Exception as e:
            email_status = f"Email failed: {e}"
        print(f"DEBUG: {email_status}", flush=True)

        automate_context.mark_run_success(
            f"Plugins: {len(plugin_breps)} | Core: {len(core_breps)} | Filtration: {len(filtration_breps)} | Pollution: {len(pollution_breps)} | {sheets_status} | {email_status}"
        )

    except Exception as e:
        print(f"EXCEPTION: {e}", flush=True)
        traceback.print_exc()
        automate_context.mark_run_failed(f"Function crashed: {e}")


if __name__ == "__main__":
    execute_automate_function(automate_function, FunctionInputs)
